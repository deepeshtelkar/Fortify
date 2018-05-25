#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import os
import csv
import time
import requests
import openpyxl as xl 


class get_urls:
	"Class to capture streaming data in real time from YAHOO"

	# -- constructor --
	def __init__(self):
		self.urls = [
			'https://jenkins.optum.com/bff/job/API-Gateway/job/API-Gateway-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/Fox-UI/job/Fox-UI-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/Fox-Claims/job/Fox-Claims-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/MDM-Mediator/job/MDM-Mediator-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/BPM-Mediator/job/BPM-Mediator-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/FOX-Process/job/Fox-Process-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/Fox-Common/job/Fox-Common-Fortify/lastSuccessfulBuild/console',
			'https://jenkins.optum.com/bff/job/Member-Validation/job/Member-Validation-Fortify/lastSuccessfulBuild/console'
		]

		self.excel_file = "job_data.xlsx"
		self.output_file = "job-data.csv"
		self.titles = ['Date', 'URL', 'Critical', 'High', 'Low']

		return


	# -- get URL content --
	def get_url_content(self, url):
		result = requests.get(url)
		page = result.text

		return page


	# -- get data from page --
	def get_data_content(self, page, url):
		row = {
			'Date': "",
			'URL': url,
			'Critical': 0,
			'High': 0,
			'Low': 0
		} 

		lines = page.split("\n")

		for i in range(0, len(lines)):
			if lines[i].find("Finished at:") >= 0:
				ini = lines[i].find(":") + 1
				end = lines[i][ini:].find("T") + ini
				tmpdate = lines[i][ini:end].strip()
				tmp = tmpdate.split("-")
				row['Date'] = tmp[1]+"/"+tmp[2]+"/"+tmp[0]

			if lines[i].find("Scan Results") >= 0:
				ini = lines[i+2].find(":") + 1
				row['Critical'] = lines[i+2][ini:].strip()
				ini = lines[i+3].find(":") + 1
				row['High'] = lines[i+3][ini:].strip()
				ini = lines[i+4].find(":") + 1
				row['Low'] = lines[i+4][ini:].strip()

		return row


	# -- write data on csv --
	def write_csv(self, csv_file, data):
		fp = open(csv_file, 'w')
		writer = csv.DictWriter(fp, fieldnames=self.titles, extrasaction='ignore', delimiter=',', quoting=csv.QUOTE_NONNUMERIC)
		writer.writeheader()
		for i in data:
			writer.writerow(i)
		fp.close()

		return


	# -- add results to excel file --
	def write_excel(self, excel_file, data):
		today = time.strftime("%m/%d/%Y", time.localtime())

		# -- open excel_file for reading to get data cells --
		wb = xl.load_workbook(excel_file, data_only=True)
		sheet = wb['Sheet1']

		ini_row, cols = self.get_columns(sheet)

		# -- reopen excel_file for writing --
		wb = xl.load_workbook(excel_file)
		sheet = wb['Sheet1']

		sheet.cell(row=ini_row, column=cols['URL']).value = today
		sheet.cell(row=ini_row, column=cols['URL']+1).value = "URL"

		r = 1
		for i in data:
			url = sheet.cell(ini_row+r, column=cols['URL']).value

			for n in data:
				if url == n['URL']:
					sheet.cell(row=ini_row+r, column=cols['URL']).value = n['Critical']
					sheet.cell(row=ini_row+r+1, column=cols['URL']).value = n['High']
					sheet.cell(row=ini_row+r+2, column=cols['URL']).value = n['Low']
					sheet.cell(row=ini_row+r, column=cols['URL']+1).value = n['URL']

			r += 4

		wb.save(excel_file)

		return


	# -- get columns from excel sheet --
	def get_columns(self, sheet):
		cols = {}
		ini_row = 0

		# -- find begin of data --
		for r in range(1, 50):
			if sheet.cell(row=r, column=1).value is not None \
			and ini_row == 0:
				ini_row = r

		for c in range(1, 50):
			if sheet.cell(row=ini_row, column=c).value is not None:
				value = sheet.cell(row=ini_row, column=c).value
				cols[value] = c

		return ini_row, cols



######################################################################################################################
## MAIN


gu = get_urls()

if __name__ == "__main__":
	data = []

	for i in gu.urls:
		page = gu.get_url_content(i)
		row = gu.get_data_content(page, i)
		data.append(row)

	#gu.write_csv(gu.output_file, data)
	gu.write_excel(gu.excel_file, data)


